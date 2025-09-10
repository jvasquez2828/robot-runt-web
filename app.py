# -*- coding: utf-8 -*-
# ==================================================================================
# === APLICACIÓN WEB FLASK PARA DESPLIEGUE EN SERVIDOR ===
# ==================================================================================
import time
import pandas as pd
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from twocaptcha import TwoCaptcha
import base64
from PIL import Image, ImageEnhance
import io
import os
from datetime import datetime
import asyncio
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from flask import Flask, render_template, Response, stream_with_context, send_from_directory
import threading
import json

# --- CONFIGURACIÓN DESDE VARIABLES DE ENTORNO ---
# Estas llaves ahora se leerán de forma segura desde el servidor.
API_KEY_2CAPTCHA = os.environ.get("API_KEY_2CAPTCHA")
GOOGLE_CREDENTIALS_JSON_STR = os.environ.get("GOOGLE_CREDENTIALS_JSON")

URL_CONSULTA = "https://portalpublico.runt.gov.co/#/consulta-vehiculo/consulta/consulta-ciudadana"
GOOGLE_SHEET_NAME = "Vehiculos a Consultar RUNT"
MAX_RETRIES = 3
CONCURRENCY_LIMIT = 4
# Para el servidor, siempre se ejecutará en modo headless (segundo plano).
HEADLESS_MODE = True

# --- INICIALIZACIÓN DE FLASK ---
app = Flask(__name__)

# --- LÓGICA DEL ROBOT (sin cambios, ya funciona) ---
async def consultar_vehiculo(page, placa, num_doc):
    captcha_id = None
    try:
        await page.goto(URL_CONSULTA, wait_until='domcontentloaded', timeout=15000)
        await page.fill("//input[@formcontrolname='placa']", placa)
        await page.click("//mat-select[@formcontrolname='tipoDocumento']")
        await page.click("//mat-option//span[contains(text(), 'NIT')]")
        await page.fill("//*[@id='mat-input-1']", str(num_doc), timeout=8000)
        
        solver = TwoCaptcha(API_KEY_2CAPTCHA)
        captcha_img_element = page.locator("xpath=//img[contains(@src, 'data:image/png')]")
        screenshot_bytes = await captcha_img_element.screenshot()
        
        image = Image.open(io.BytesIO(screenshot_bytes))
        image = image.convert('L')
        enhancer = ImageEnhance.Contrast(image)
        image = enhancer.enhance(2.5)
        enhancer = ImageEnhance.Sharpness(image)
        image = enhancer.enhance(2.0)
        threshold = 150 
        image = image.point(lambda p: 0 if p < threshold else 255)
        
        buffer = io.BytesIO()
        image.save(buffer, format="PNG")
        base64_captcha = base64.b64encode(buffer.getvalue()).decode('utf-8')
        
        resultado = await asyncio.to_thread(
            solver.normal, base64_captcha, caseSensitive=1, min_len=5, max_len=5, numeric=4, hint_text='The code is 5 characters, case-sensitive.'
        )
        codigo_captcha = resultado['code']
        captcha_id = resultado['captchaId']
        
        await page.fill("//*[@id='mat-input-2']", codigo_captcha)
        await page.press("//*[@id='mat-input-2']", "Enter")

        try:
            error_locator = page.locator("xpath=//div[contains(text(), 'código de verificación es incorrecto')]")
            await error_locator.wait_for(timeout=3500) 
            error_text = await error_locator.inner_text()
            if captcha_id:
                await asyncio.to_thread(solver.report, captcha_id, False)
            raise Exception("Error de CAPTCHA.")
        except PlaywrightTimeoutError:
            pass

        await page.wait_for_selector("text=Información general del vehículo", timeout=12000)

        soat_header_locator = page.locator("xpath=//mat-expansion-panel-header[contains(., 'Póliza SOAT')]")
        await soat_header_locator.click()
        await asyncio.sleep(0.2)
        
        estado_locator = page.locator(f"xpath=//*[@id='cdk-accordion-child-1']/div/mat-card-content/div/mat-table/mat-row[1]/mat-cell[7]")
        texto_completo_soat = (await estado_locator.inner_text(timeout=4000)).strip().lower()
        soat_info = 'Vigente' if 'vigente' in texto_completo_soat and 'no vigente' not in texto_completo_soat else 'No Vigente'
        
        limitaciones_header_locator = page.locator("xpath=//mat-expansion-panel-header[contains(., 'Limitaciones a la Propiedad')]")
        await limitaciones_header_locator.click()
        await asyncio.sleep(0.2)
        
        limitaciones_content_locator = limitaciones_header_locator.locator("xpath=./ancestor::mat-expansion-panel//div[contains(@class, 'mat-expansion-panel-content')]")
        limitaciones_info = (await limitaciones_content_locator.inner_text(timeout=4000)).strip().replace('\n', ' ')
        
        return {"SOAT": soat_info, "Limitaciones": limitaciones_info, "error": None}

    except Exception as e:
        error_msg = str(e).split('\n')[0]
        return {"SOAT": "Error", "Limitaciones": "Error", "error": error_msg}

async def process_vehicle_with_retries(browser, placa, num_doc, progress_queue, semaphore):
    async with semaphore:
        for attempt in range(MAX_RETRIES):
            context = await browser.new_context(user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.102 Safari/537.36")
            page = await context.new_page()
            
            async def handle_route(route):
                if route.request.resource_type in ["stylesheet", "font", "media", "image"] and not route.request.url.startswith("data:image"):
                    await route.abort()
                else:
                    await route.continue_()
            await page.route("**/*", handle_route)

            resultado = await consultar_vehiculo(page, placa, num_doc)
            await context.close()

            if resultado['error'] is None:
                progress_queue.put(1)
                return {'placa': placa, **resultado}
            else:
                if attempt < MAX_RETRIES - 1:
                    await asyncio.sleep(1.5)

        progress_queue.put(1)
        return {'placa': placa, **resultado}

def run_scraper_process(progress_queue):
    asyncio.run(main_scraper(progress_queue))

async def main_scraper(progress_queue):
    try:
        if not API_KEY_2CAPTCHA or not GOOGLE_CREDENTIALS_JSON_STR:
            raise ValueError("Las variables de entorno no están configuradas en el servidor.")
            
        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        # Carga las credenciales desde la variable de entorno (string)
        creds_dict = json.loads(GOOGLE_CREDENTIALS_JSON_STR)
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        client = gspread.authorize(creds)
        sheet = client.open(GOOGLE_SHEET_NAME).sheet1
        data = sheet.get_all_records()
        df_entrada = pd.DataFrame(data)
        
        total_count = len(df_entrada)
        progress_queue.put({'total': total_count})
        
    except Exception as e:
        progress_queue.put({'error': f"Error al leer Google Sheets: {e}"})
        return

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=HEADLESS_MODE)
        semaphore = asyncio.Semaphore(CONCURRENCY_LIMIT)
        tasks = []
        for _, fila in df_entrada.iterrows():
            tasks.append(process_vehicle_with_retries(browser, fila['placa'], str(fila['numero_documento']), progress_queue, semaphore))
        
        lista_resultados = await asyncio.gather(*tasks)
        await browser.close()

    df_resultados = pd.DataFrame(lista_resultados)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    # Guardamos el archivo en una carpeta temporal 'tmp' que es estándar en muchos servidores.
    if not os.path.exists('tmp'):
        os.makedirs('tmp')
    filename = os.path.join('tmp', f'resultados_consulta_{timestamp}.xlsx')
    df_resultados.to_excel(filename, index=False)
    
    try:
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        wb = load_workbook(filename)
        ws = wb.active
        for row in range(2, ws.max_row + 1):
            celda_soat = ws.cell(row=row, column=2)
            if celda_soat.value and 'Vigente' in str(celda_soat.value):
                 celda_soat.fill = green_fill
            elif celda_soat.value and celda_soat.value != 'Error':
                celda_soat.fill = red_fill
        wb.save(filename)
    except Exception:
        pass
        
    progress_queue.put({'done': os.path.basename(filename)})

@app.route('/')
def index():
    return render_template('index.html')

scraper_thread = None
progress_queue = None

@app.route('/run')
def run_scraper():
    global scraper_thread, progress_queue
    if scraper_thread and scraper_thread.is_alive():
        return "El proceso ya está en ejecución.", 400
    
    from queue import Queue
    progress_queue = Queue()
    
    scraper_thread = threading.Thread(target=run_scraper_process, args=(progress_queue,))
    scraper_thread.start()
    return "Proceso iniciado.", 200

@app.route('/progress')
def progress():
    def generate():
        global progress_queue
        if not progress_queue: return
        completed_count = 0
        total_count = 1
        
        while True:
            try:
                message = progress_queue.get()
                if isinstance(message, dict):
                    if 'total' in message: total_count = message['total']
                    elif 'error' in message:
                        yield f"data: {json.dumps({'status': message['error']})}\n\n"; break
                    elif 'done' in message:
                        yield f"data: {json.dumps({'status': '¡Proceso completado!', 'progress': 100, 'file': message['done']})}\n\n"; break
                else:
                    completed_count += message
                    progress = (completed_count / total_count) * 100
                    yield f"data: {json.dumps({'status': f'Procesando {completed_count}/{total_count}...', 'progress': progress})}\n\n"
                time.sleep(0.1)
            except Exception: break
                
    return Response(stream_with_context(generate()), mimetype='text/event-stream')

@app.route('/download/<filename>')
def download_file(filename):
    # Los archivos se descargan desde la carpeta temporal 'tmp'.
    return send_from_directory('tmp', filename, as_attachment=True)

# Esta parte ya no se usa para iniciar el servidor en producción,
# pero es útil para pruebas locales. El servidor usará 'gunicorn'.
if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')








